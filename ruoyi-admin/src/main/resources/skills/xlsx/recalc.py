#!/usr/bin/env python3
"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
from openpyxl import load_workbook


def get_soffice_path():
    """Find LibreOffice soffice executable path for the current OS"""
    system = platform.system()

    if system == 'Windows':
        # Common LibreOffice installation paths on Windows
        possible_paths = [
            Path(os.environ.get('PROGRAMFILES', 'C:\\Program Files')) / 'LibreOffice' / 'program' / 'soffice.exe',
            Path(os.environ.get('PROGRAMFILES(X86)', 'C:\\Program Files (x86)')) / 'LibreOffice' / 'program' / 'soffice.exe',
            Path(os.path.expanduser('~')) / 'AppData' / 'Local' / 'LibreOffice' / 'program' / 'soffice.exe',
        ]

        for path in possible_paths:
            if path.exists():
                return str(path)

        # Try to find it using where command
        try:
            result = subprocess.run(['where', 'soffice.exe'], capture_output=True, text=True, timeout=5)
            if result.returncode == 0:
                return result.stdout.strip().split('\n')[0]
        except Exception:
            pass

        return None
    else:
        # For Linux and macOS, soffice is usually in PATH
        return 'soffice'


def setup_libreoffice_macro():
    """Setup LibreOffice macro for recalculation if not already configured"""
    system = platform.system()

    if system == 'Darwin':
        macro_dir = os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    elif system == 'Windows':
        # Windows path for LibreOffice config
        appdata = os.path.expanduser('~\\AppData\\Roaming\\LibreOffice\\4\\user\\basic\\Standard')
        macro_dir = appdata
    else:
        # Linux
        macro_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')

    macro_file = os.path.join(macro_dir, 'Module1.xba')

    if os.path.exists(macro_file):
        try:
            with open(macro_file, 'r') as f:
                if 'RecalculateAndSave' in f.read():
                    return True
        except Exception:
            pass

    if not os.path.exists(macro_dir):
        soffice_path = get_soffice_path()
        if not soffice_path:
            return False

        try:
            subprocess.run([soffice_path, '--headless', '--terminate_after_init'],
                          capture_output=True, timeout=10)
        except Exception:
            pass

        os.makedirs(macro_dir, exist_ok=True)
    
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''
    
    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """
    Recalculate formulas in Excel file and report any errors

    Args:
        filename: Path to Excel file
        timeout: Maximum time to wait for recalculation (seconds)

    Returns:
        dict with error locations and counts
    """
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {'error': 'Failed to setup LibreOffice macro'}

    soffice_path = get_soffice_path()
    if not soffice_path:
        return {'error': 'LibreOffice not found. Please install LibreOffice.'}

    cmd = [
        soffice_path, '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]

    system = platform.system()

    # Handle timeout for different operating systems
    if system == 'Windows':
        # Windows: use taskkill as fallback, but rely on subprocess timeout
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        except subprocess.TimeoutExpired:
            return {'error': f'LibreOffice recalculation timed out after {timeout} seconds'}
    elif system == 'Linux':
        # Linux: use timeout command
        timeout_cmd = 'timeout'
        cmd = [timeout_cmd, str(timeout)] + cmd
        result = subprocess.run(cmd, capture_output=True, text=True)
    elif system == 'Darwin':
        # macOS: try gtimeout first, fallback to timeout handling
        timeout_cmd = None
        try:
            subprocess.run(['gtimeout', '--version'], capture_output=True, timeout=1, check=False)
            timeout_cmd = 'gtimeout'
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

        if timeout_cmd:
            cmd = [timeout_cmd, str(timeout)] + cmd
            result = subprocess.run(cmd, capture_output=True, text=True)
        else:
            try:
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
            except subprocess.TimeoutExpired:
                return {'error': f'LibreOffice recalculation timed out after {timeout} seconds'}
    else:
        result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0 and result.returncode != 124:  # 124 is timeout exit code
        error_msg = result.stderr or 'Unknown error during recalculation'
        if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
            return {'error': 'LibreOffice macro not configured properly. Error: ' + error_msg}
        else:
            return {'error': error_msg}
    
    # Check for Excel errors in the recalculated file - scan ALL cells
    try:
        wb = load_workbook(filename, data_only=True)
        
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check ALL rows and columns - no limits
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break
        
        wb.close()
        
        # Build result summary
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {}
        }
        
        # Add non-empty error categories
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]  # Show up to 20 locations
                }
        
        # Add formula count for context - also check ALL cells
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_formulas.close()
        
        result['total_formulas'] = formula_count
        
        return result
        
    except Exception as e:
        return {'error': str(e)}


def get_project_root():
    """Get project root directory (ruoyi-ai-v3)"""
    current = os.path.dirname(os.path.abspath(__file__))
    # Traverse up 8 levels from xlsx/recalc.py to project root
    for _ in range(8):
        current = os.path.dirname(current)
    return current


def create_workbook(output_path, sheets_config):
    """
    Create a new Excel workbook with specified sheets and data

    Args:
        output_path: Path where the workbook will be saved
        sheets_config: Dict with sheet configurations
            Example: {
                'Sheet1': {
                    'data': [
                        ['Header1', 'Header2'],
                        ['Value1', '=B2*2']
                    ],
                    'formulas': False  # Whether sheet contains formulas
                }
            }

    Returns:
        dict with success/error status
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for sheet_name, config in sheets_config.items():
            ws = wb.create_sheet(sheet_name)
            data = config.get('data', [])

            for row_idx, row_data in enumerate(data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    # Make header row bold
                    if row_idx == 1:
                        cell.font = Font(bold=True)

        wb.save(output_path)
        return {'status': 'success', 'message': f'Workbook created at {output_path}'}

    except Exception as e:
        return {'status': 'error', 'message': str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python recalc.py <excel_file> [timeout_seconds]     # Recalculate formulas")
        print("  python recalc.py --create <output_file> <data_json>  # Create new workbook")
        print("\nRecalculate formulas:")
        print("  Recalculates all formulas in an Excel file using LibreOffice")
        print("  Returns JSON with error details")
        print("\nCreate workbook:")
        print("  data_json format: '{\"Sheet1\": {\"data\": [[\"A\", \"B\"], [1, 2]]}}'")
        sys.exit(1)

    if sys.argv[1] == '--create':
        if len(sys.argv) < 4:
            print("Error: --create requires output_file and data_json")
            sys.exit(1)
        output_file = sys.argv[2]
        # If relative path, place in project root
        if not os.path.isabs(output_file):
            project_root = get_project_root()
            output_file = os.path.join(project_root, output_file)
        data_json = sys.argv[3]
        try:
            sheets_config = json.loads(data_json)
            result = create_workbook(output_file, sheets_config)
            print(json.dumps(result, indent=2))
        except json.JSONDecodeError as e:
            print(json.dumps({'status': 'error', 'message': f'Invalid JSON: {str(e)}'}, indent=2))
    else:
        filename = sys.argv[1]
        timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
        result = recalc(filename, timeout)
        print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()